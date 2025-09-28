import os
import uuid
import pandas as pd
import logging
import time
from pathlib import Path
from flask import Flask, request, jsonify, send_file, render_template
from flask_socketio import SocketIO, emit, join_room
from werkzeug.utils import secure_filename
from threading import Thread
from text_extractor import PDFTextExtractor
from invoice_parser import InvoiceParser
from matrix_processor import UPSMatrixProcessor
import fitz  # PyMuPDF
from PIL import Image
import io
from typing import List, Tuple, Dict
import re

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['MAX_CONTENT_LENGTH'] = 5000000000 * 1024 * 1024  # 50MB max file size
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'

# Initialize SocketIO
socketio = SocketIO(app, cors_allowed_origins="*", logger=False, engineio_logger=False)

# Create directories
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'pdf'}

# Store active sessions
active_sessions = {}

class DirectPDFExtractor:
    """Direct PDF text extractor for extracting the 5 key fields"""
    
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.doc = fitz.open(pdf_path)
    
    def extract_page_data(self, page_num: int) -> Tuple[Image.Image, List[str], List[List[int]]]:
        """Extract image, words, and bounding boxes from a PDF page"""
        page = self.doc[page_num]
        
        # Convert page to image
        mat = fitz.Matrix(2, 2)  # 2x zoom for better quality
        pix = page.get_pixmap(matrix=mat)
        img_data = pix.tobytes("png")
        image = Image.open(io.BytesIO(img_data))
        
        # Extract text with coordinates
        text_dict = page.get_text("dict")
        words, boxes = self._extract_words_and_boxes(text_dict)
        
        return image, words, boxes
    
    def _extract_words_and_boxes(self, text_dict: dict) -> Tuple[List[str], List[List[int]]]:
        """Extract words and bounding boxes from text dictionary"""
        words = []
        boxes = []
        
        for block in text_dict["blocks"]:
            if "lines" in block:
                for line in block["lines"]:
                    for span in line["spans"]:
                        text = span["text"].strip()
                        if text:
                            bbox = span["bbox"]
                            words.append(text)
                            boxes.append([int(bbox[0]), int(bbox[1]), int(bbox[2]), int(bbox[3])])
        
        return words, boxes
    
    def get_total_pages(self) -> int:
        """Get total number of pages in PDF"""
        return len(self.doc)
    
    def is_empty_page(self, page_num: int) -> bool:
        """Check if page is empty or has minimal content"""
        _, words, _ = self.extract_page_data(page_num)
        return len(words) < 5
    
    def close(self):
        """Close the PDF document"""
        self.doc.close()

class DirectInvoiceParser:
    """Direct invoice parser focused on extracting the 5 key fields accurately"""
    
    def __init__(self):
        pass
    
    def extract_direct_fields(self, pdf_path: str) -> List[Dict[str, str]]:
        """Extract the 5 key fields directly from PDF"""
        extractor = DirectPDFExtractor(pdf_path)
        all_extracted_data = []
        
        try:
            total_pages = extractor.get_total_pages()
            logger.info(f"Direct extraction: Processing {total_pages} pages")
            
            for page_num in range(total_pages):
                if extractor.is_empty_page(page_num):
                    continue
                
                image, words, boxes = extractor.extract_page_data(page_num)
                page_text = ' '.join(words)
                
                # Extract shipments from this page
                page_shipments = self._extract_shipments_from_page(page_text, page_num + 1)
                all_extracted_data.extend(page_shipments)
                
        except Exception as e:
            logger.error(f"Error in direct extraction: {e}")
        finally:
            extractor.close()
        
        logger.info(f"Direct extraction completed: {len(all_extracted_data)} shipments found")
        return all_extracted_data
    
    def _extract_shipments_from_page(self, text: str, page_num: int) -> List[Dict[str, str]]:
        """Extract shipments with 5 key fields from a single page"""
        shipments = []
        
        # Find all tracking numbers as shipment boundaries
        tracking_pattern = r'(1Z[A-Z0-9]{16})'
        tracking_matches = list(re.finditer(tracking_pattern, text))
        
        for i, match in enumerate(tracking_matches):
            tracking_number = match.group(1)
            
            # Define text block for this shipment
            start_pos = match.start()
            if i + 1 < len(tracking_matches):
                end_pos = tracking_matches[i + 1].start()
            else:
                end_pos = len(text)
            
            shipment_block = text[start_pos:end_pos]
            
            # Extract the 5 key fields from this block
            extracted_data = self._extract_five_fields(shipment_block)
            extracted_data['tracking_number'] = tracking_number
            extracted_data['page_number'] = page_num
            
            shipments.append(extracted_data)
        
        return shipments
    
    def _extract_five_fields(self, block: str) -> Dict[str, str]:
        """Extract the 5 key fields from a shipment block with improved sender/receiver logic"""
        data = {
            'tracking_number': '',
            'sender_name': '',
            'sender_address': '',
            'receiver_name': '',
            'receiver_address': ''
        }
        
        # Clean the block
        block = ' '.join(block.split())
        
        # Method 1: Look for explicit "Sender:" and "Receiver:" patterns
        sender_pattern = r'Sender\s*:\s*([A-Z][A-Za-z\s]{2,40}?)\s+(\d+[^:]*?[A-Z]{2}\s+\d{5}(?:-\d{4})?)'
        sender_match = re.search(sender_pattern, block, re.IGNORECASE)
        
        if sender_match:
            potential_name = sender_match.group(1).strip()
            potential_address = sender_match.group(2).strip()
            
            # Clean sender name - remove any trailing address components
            clean_name = re.sub(r'\s+\d+.*', '', potential_name).strip()
            if len(clean_name) >= 4 and not re.match(r'^\d', clean_name):
                data['sender_name'] = clean_name
                data['sender_address'] = potential_address
        
        receiver_pattern = r'Receiver\s*:\s*([A-Z][A-Za-z\s]{2,40}?)\s+(\d+[^:]*?[A-Z]{2}\s+\d{5}(?:-\d{4})?)'
        receiver_match = re.search(receiver_pattern, block, re.IGNORECASE)
        
        if receiver_match:
            potential_name = receiver_match.group(1).strip()
            potential_address = receiver_match.group(2).strip()
            
            # Clean receiver name - remove any trailing address components
            clean_name = re.sub(r'\s+\d+.*', '', potential_name).strip()
            if len(clean_name) >= 4 and not re.match(r'^\d', clean_name):
                data['receiver_name'] = clean_name
                data['receiver_address'] = potential_address
        
        # Method 2: Improved pattern-based extraction if explicit patterns didn't work
        if not data['sender_name'] or not data['receiver_name']:
            
            # Split the text into logical sections
            # First, find key markers to understand the structure
            parts = re.split(r'(UserID:|1st ref:|2nd ref:|Sender\s*:|Receiver\s*:)', block, flags=re.IGNORECASE)
            
            # Find all potential names and addresses more carefully
            name_pattern = r'\b([A-Z]{2,}(?:\s+[A-Z]{2,})*)\b'
            address_pattern = r'(\d+\s+[A-Z0-9][^:]*?[A-Z]{2}\s+\d{5}(?:-\d{4})?)'
            
            all_names = []
            all_addresses = []
            
            for name_match in re.finditer(name_pattern, block):
                name = name_match.group(1)
                # Filter out non-names more strictly
                if not re.match(r'^(DELIVERY|SERVICE|INVOICE|CUSTOMER|WEIGHT|RESIDENTIAL|SURCHARGE|FUEL|DIMENSIONS|TOTAL|USERIDS?|SENDER|RECEIVER|GROUND|NEXT|DAY|AIR|TRACKING|NUMBER|ACCOUNT|PAGE|VENTURE|CT|KY|NV|AVE|STREET|ST|ROAD|RD|DRIVE|DR|LANE|LN|COURT|BLVD|BOULEVARD)(\s+\w+)*$', name, re.IGNORECASE):
                    # Additional check: name should not contain numbers or be too long
                    if not re.search(r'\d', name) and 4 <= len(name) <= 40:
                        all_names.append((name, name_match.start()))
            
            for addr_match in re.finditer(address_pattern, block, re.IGNORECASE):
                address = addr_match.group(1)
                # Clean address - remove service types and other non-address content
                cleaned_addr = re.sub(r'\s+(Ground|Air|Next|Day|Residential|Commercial)\s+', ' ', address, flags=re.IGNORECASE)
                cleaned_addr = re.sub(r'\s+(Service|Surcharge|Weight|Total)\b.*', '', cleaned_addr, flags=re.IGNORECASE)
                cleaned_addr = ' '.join(cleaned_addr.split())
                
                if len(cleaned_addr) >= 10:  # Minimum address length
                    all_addresses.append((cleaned_addr, addr_match.start()))
            
            # Now try to match names and addresses by position and context
            sender_section_end = -1
            receiver_section_start = float('inf')
            
            # Find where sender section ends and receiver section starts
            sender_match = re.search(r'Sender\s*:', block, re.IGNORECASE)
            receiver_match = re.search(r'Receiver\s*:', block, re.IGNORECASE)
            
            if sender_match:
                sender_section_end = sender_match.end()
            if receiver_match:
                receiver_section_start = receiver_match.start()
            
            # Assign names and addresses based on position
            if not data['sender_name'] and all_names:
                # Look for sender name after "Sender:" or in first half
                for name, pos in all_names:
                    if sender_section_end > 0 and pos > sender_section_end and pos < receiver_section_start:
                        data['sender_name'] = name
                        break
                
                # Fallback: first valid name if no positional match
                if not data['sender_name']:
                    data['sender_name'] = all_names[0][0]
            
            if not data['receiver_name'] and len(all_names) > 1:
                # Look for receiver name after "Receiver:"
                for name, pos in all_names:
                    if pos > receiver_section_start:
                        data['receiver_name'] = name
                        break
                
                # Fallback: last valid name if no positional match
                if not data['receiver_name']:
                    data['receiver_name'] = all_names[-1][0]
            
            # Assign addresses similarly
            if not data['sender_address'] and all_addresses:
                for address, pos in all_addresses:
                    if sender_section_end > 0 and pos > sender_section_end and pos < receiver_section_start:
                        data['sender_address'] = address
                        break
                
                if not data['sender_address']:
                    data['sender_address'] = all_addresses[0][0]
            
            if not data['receiver_address'] and len(all_addresses) > 1:
                for address, pos in all_addresses:
                    if pos > receiver_section_start:
                        data['receiver_address'] = address
                        break
                
                if not data['receiver_address']:
                    data['receiver_address'] = all_addresses[-1][0]
        
        # Method 3: Final cleanup and validation
        # Clean sender name to ensure it doesn't contain address parts
        if data['sender_name']:
            # Remove any numeric parts that might be addresses
            cleaned_name = re.sub(r'\s+\d+\s+.*', '', data['sender_name']).strip()
            # Remove common address words
            cleaned_name = re.sub(r'\s+(CT|COURT|AVE|AVENUE|ST|STREET|RD|ROAD|DR|DRIVE|BLVD|BOULEVARD|LN|LANE)\b.*', '', cleaned_name, flags=re.IGNORECASE).strip()
            # Remove state abbreviations 
            cleaned_name = re.sub(r'\s+[A-Z]{2}\s+\d', '', cleaned_name).strip()
            
            if len(cleaned_name) >= 4:
                data['sender_name'] = cleaned_name
            else:
                data['sender_name'] = ''
        
        # Clean receiver name similarly
        if data['receiver_name']:
            cleaned_name = re.sub(r'\s+\d+\s+.*', '', data['receiver_name']).strip()
            cleaned_name = re.sub(r'\s+(CT|COURT|AVE|AVENUE|ST|STREET|RD|ROAD|DR|DRIVE|BLVD|BOULEVARD|LN|LANE)\b.*', '', cleaned_name, flags=re.IGNORECASE).strip()
            cleaned_name = re.sub(r'\s+[A-Z]{2}\s+\d', '', cleaned_name).strip()
            
            if len(cleaned_name) >= 4:
                data['receiver_name'] = cleaned_name
            else:
                data['receiver_name'] = ''
        
        return data

    def _clean_name(self, name: str) -> str:
        """Enhanced name cleaning with better validation"""
        if not name:
            return ''
        
        # Remove common non-name words and address components
        name = re.sub(r'\s+(Customer|Weight|Residential|Surcharge|Fuel|Next|Day|Air|Ground|Total|Service|Tracking|Number)', '', name, flags=re.IGNORECASE)
        
        # Remove address components
        name = re.sub(r'\s+(CT|COURT|AVE|AVENUE|ST|STREET|RD|ROAD|DR|DRIVE|BLVD|BOULEVARD|LN|LANE|ZIP|CODE|ZONE)\b.*', '', name, flags=re.IGNORECASE)
        
        # Remove any trailing numbers that might be addresses
        name = re.sub(r'\s+\d+.*', '', name)
        
        # Remove state abbreviations and zip codes
        name = re.sub(r'\s+[A-Z]{2}\s+\d{5}.*', '', name)
        
        # Normalize whitespace
        name = ' '.join(name.split())
        
        # Final validation - name should be mostly letters and reasonable length
        if len(name) < 4 or len(name) > 50:
            return ''
        
        # Should not start with numbers or common non-name words
        if re.match(r'^(\d|Ground|Air|Service|Residential)', name, re.IGNORECASE):
            return ''
        
        return name.strip()

    def _clean_address(self, address: str) -> str:
        """Enhanced address cleaning with better service type removal"""
        if not address:
            return ''
        
        # Remove prices and weights at the beginning
        address = re.sub(r'^\s*\d+\.\d+\s*-?\d*\.\d*\s*\d+\.\d+\s*', '', address)
        
        # Remove service types that might be mixed in
        address = re.sub(r'\s+(Ground|Air|Next|Day)\s+(Residential|Commercial)?\s*', ' ', address, flags=re.IGNORECASE)
        
        # Remove common invoice terms
        address = re.sub(r'\s+(Customer Weight|Residential Surcharge|Fuel Surcharge|Total|1st ref|UserID|Sender).*', '', address, flags=re.IGNORECASE)
        
        # Remove tracking numbers that might have been included
        address = re.sub(r'\s+1Z[A-Z0-9]{16}\s+', ' ', address)
        
        # Clean up extra spaces and normalize
        address = ' '.join(address.split())
        
        # Final validation - address should have reasonable length and contain numbers
        if len(address) < 10 or not re.search(r'\d', address):
            return ''
        
        return address.strip()

def emit_progress(session_id, data):
    """Helper function to emit progress with proper error handling"""
    try:
        socketio.emit('progress_update', data, room=session_id)
        time.sleep(0.05)
    except Exception as e:
        logger.error(f"Failed to emit progress to session {session_id}: {e}")

def emit_completion(session_id, data):
    """Helper function to emit completion with proper error handling"""
    try:
        socketio.emit('processing_complete', data, room=session_id)
        time.sleep(0.1)
    except Exception as e:
        logger.error(f"Failed to emit completion to session {session_id}: {e}")

def emit_error(session_id, data):
    """Helper function to emit error with proper error handling"""
    try:
        socketio.emit('processing_error', data, room=session_id)
        time.sleep(0.1)
    except Exception as e:
        logger.error(f"Failed to emit error to session {session_id}: {e}")

def process_invoice_with_progress(pdf_path, output_path, session_id):
    """Enhanced invoice processing with direct field extraction and replacement"""
    try:
        logger.info(f"=== STARTING ENHANCED PROCESSING WITH DIRECT EXTRACTION for session: {session_id} ===")
        
        # Check if session is active
        if session_id not in active_sessions:
            logger.error(f"Session {session_id} not found in active sessions")
            return 0, 0
        
        # Step 1: Original matrix-based extraction (existing functionality)
        emit_progress(session_id, {
            'current_page': 0,
            'total_pages': 0,
            'percentage': 5,
            'status': 'Starting original matrix-based extraction...',
            'shipments_found': 0
        })
        
        extractor = PDFTextExtractor(pdf_path)
        parser = InvoiceParser()
        matrix_processor = UPSMatrixProcessor()
        
        # Extract invoice groups with existing method
        invoice_groups = extractor.extract_invoice_groups()
        logger.info(f"Found {len(invoice_groups)} invoice groups")
        
        emit_progress(session_id, {
            'current_page': 0,
            'total_pages': len(invoice_groups),
            'percentage': 15,
            'status': f'Found {len(invoice_groups)} invoice groups. Processing with matrix extraction...',
            'shipments_found': 0
        })
        
        # Process each invoice group (existing logic)
        all_shipments = []
        total_groups = len(invoice_groups)
        
        for group_idx, group in enumerate(invoice_groups):
            logger.info(f"=== PROCESSING INVOICE GROUP {group_idx + 1} of {total_groups} ===")
            
            # Get invoice number for proper heading
            invoice_header = group.get('invoice_header', {})
            invoice_number = invoice_header.get('invoice_number', f'Invoice_{group_idx + 1}')
            
            # Emit progress for current group
            progress_data = {
                'current_page': group_idx + 1,
                'total_pages': total_groups,
                'percentage': int(15 + (group_idx / total_groups) * 40),
                'status': f'Processing invoice {invoice_number} ({group_idx + 1} of {total_groups})',
                'shipments_found': len(all_shipments),
                'current_invoice': invoice_number
            }
            emit_progress(session_id, progress_data)
            
            try:
                # Process pages in this invoice group (existing logic)
                group_shipments = []
                
                for page_num in group['pages']:
                    if extractor.is_empty_page(page_num):
                        logger.info(f"Page {page_num + 1} is empty, skipping")
                        continue
                    
                    logger.info(f"Processing page {page_num + 1} with matrix extraction")
                    image, words, boxes = extractor.extract_page_data(page_num)
                    
                    # Parse using existing matrix-based extraction
                    page_shipments = parser.parse_invoice(image, words, boxes)
                    
                    if page_shipments:
                        logger.info(f"Found {len(page_shipments)} shipments on page {page_num + 1}")
                        for shipment in page_shipments:
                            shipment['page_number'] = page_num + 1
                            shipment['invoice_group'] = group_idx + 1
                            
                            # Add invoice header data
                            for key, value in invoice_header.items():
                                if not shipment.get(key) and value:
                                    shipment[key] = value
                            
                            # Ensure invoice number is properly set
                            if not shipment.get('invoice_number'):
                                shipment['invoice_number'] = invoice_number
                        
                        group_shipments.extend(page_shipments)
                
                logger.info(f"Invoice group {invoice_number} completed with {len(group_shipments)} shipments")
                all_shipments.extend(group_shipments)
                
            except Exception as e:
                logger.error(f"Error processing invoice group {invoice_number}: {e}")
                continue
        
        extractor.close()
        logger.info(f"=== MATRIX EXTRACTION COMPLETED. Total shipments: {len(all_shipments)} ===")
        
        # Step 2: Direct field extraction for the 5 key fields
        emit_progress(session_id, {
            'current_page': total_groups,
            'total_pages': total_groups,
            'percentage': 60,
            'status': 'Starting direct extraction of key fields (tracking, sender, receiver)...',
            'shipments_found': len(all_shipments)
        })
        
        logger.info("=== STARTING DIRECT FIELD EXTRACTION ===")
        direct_parser = DirectInvoiceParser()
        direct_extracted_data = direct_parser.extract_direct_fields(pdf_path)
        
        logger.info(f"Direct extraction completed: {len(direct_extracted_data)} shipments")
        
        # Step 3: Merge the direct extracted fields with matrix data
        emit_progress(session_id, {
            'current_page': total_groups,
            'total_pages': total_groups,
            'percentage': 75,
            'status': 'Merging direct extracted fields with matrix data...',
            'shipments_found': len(all_shipments)
        })
        
        merged_shipments = merge_direct_fields_with_matrix(all_shipments, direct_extracted_data)
        
        # Step 4: Generate Excel with enhanced data
        emit_progress(session_id, {
            'current_page': total_groups,
            'total_pages': total_groups,
            'percentage': 85,
            'status': 'Generating Excel file with corrected address fields...',
            'shipments_found': len(merged_shipments)
        })
        
        if merged_shipments:
            logger.info("Creating enhanced Excel file with corrected address fields")
            create_enhanced_matrix_excel(merged_shipments, output_path)
            logger.info(f"Enhanced Excel saved successfully: {output_path}")
            
            # Generate processing statistics
            stats = generate_processing_statistics(merged_shipments)
            stats['direct_extraction_count'] = len(direct_extracted_data)
            stats['merge_success_rate'] = calculate_merge_success_rate(all_shipments, direct_extracted_data)
            
            # Emit completion with detailed statistics
            completion_data = {
                'success': True,
                'shipment_count': len(merged_shipments),
                'invoice_count': total_groups,
                'message': f'Successfully extracted {len(merged_shipments)} shipments with corrected address fields',
                'excel_structure': 'Enhanced Matrix with Direct Field Correction',
                'statistics': stats
            }
            emit_completion(session_id, completion_data)
            
            return len(merged_shipments), total_groups
        
        # No data found
        no_data_completion = {
            'success': False,
            'shipment_count': 0,
            'invoice_count': total_groups,
            'message': 'No shipment data found in the PDF'
        }
        emit_completion(session_id, no_data_completion)
        
        return 0, total_groups
        
    except Exception as e:
        logger.error(f"Error processing PDF: {e}", exc_info=True)
        error_data = {'error': f'Processing failed: {str(e)}'}
        emit_error(session_id, error_data)
        raise e

def merge_direct_fields_with_matrix(matrix_shipments: List[Dict], direct_data: List[Dict]) -> List[Dict]:
    """Merge directly extracted fields with matrix data based on tracking numbers"""
    logger.info(f"=== MERGING DIRECT FIELDS WITH MATRIX DATA ===")
    logger.info(f"Matrix shipments: {len(matrix_shipments)}, Direct data: {len(direct_data)}")
    
    # Create lookup dictionary from direct data
    direct_lookup = {}
    for data in direct_data:
        tracking = data.get('tracking_number')
        if tracking:
            direct_lookup[tracking] = data
    
    merged_count = 0
    
    # Merge direct fields into matrix shipments
    for shipment in matrix_shipments:
        tracking = shipment.get('tracking_number')
        if tracking and tracking in direct_lookup:
            direct_fields = direct_lookup[tracking]
            
            # Replace the 5 key fields with directly extracted data
            if direct_fields.get('sender_name'):
                shipment['sender_name'] = direct_fields['sender_name']
            if direct_fields.get('sender_address'):
                shipment['sender_address'] = direct_fields['sender_address']
            if direct_fields.get('receiver_name'):
                shipment['receiver_name'] = direct_fields['receiver_name']
            if direct_fields.get('receiver_address'):
                shipment['receiver_address'] = direct_fields['receiver_address']
            
            # Add flag indicating direct extraction was used
            shipment['direct_extraction_applied'] = True
            merged_count += 1
        else:
            shipment['direct_extraction_applied'] = False
    
    logger.info(f"Successfully merged {merged_count} shipments with direct extraction data")
    return matrix_shipments

def calculate_merge_success_rate(matrix_shipments: List[Dict], direct_data: List[Dict]) -> float:
    """Calculate the success rate of merging direct data with matrix data"""
    if not matrix_shipments:
        return 0.0
    
    direct_tracking_numbers = set(d.get('tracking_number') for d in direct_data if d.get('tracking_number'))
    matrix_tracking_numbers = set(s.get('tracking_number') for s in matrix_shipments if s.get('tracking_number'))
    
    matched = len(direct_tracking_numbers.intersection(matrix_tracking_numbers))
    total = len(matrix_tracking_numbers)
    
    return (matched / total * 100) if total > 0 else 0.0

def create_enhanced_matrix_excel(shipments: list, output_path: str):
    """Create Excel file with corrected sender/receiver fields from direct extraction"""
    
    # Group shipments by invoice number
    invoice_groups = {}
    for shipment in shipments:
        inv_num = shipment.get('invoice_number', 'Unknown')
        if inv_num not in invoice_groups:
            invoice_groups[inv_num] = []
        invoice_groups[inv_num].append(shipment)
    
    # Prepare Excel data with enhanced matrix structure
    excel_rows = []
    
    # Enhanced column order with corrected address fields
    column_order = [
        # Header and identification columns
        'ROW_TYPE', 'INVOICE_GROUP_HEADER', 'SHIPMENT_INDEX', 'SHIPMENT_COUNT', 'TOTAL_SHIPMENTS',
        
        # Core invoice and shipment fields
        'invoice_number', 'tracking_number', 'account_number', 'invoice_date',
        'destination_zip', 'page_number', 'invoice_group', 'processing_type',
        
        # Shipment details
        'weight', 'zone', 'service_type', 'published_charge', 'incentive_credit', 'billed_charge',
        'shipment_date', 'pickup_date',
        
        # CORRECTED - Direct extracted address information (these will be accurate now)
        'sender_name', 'sender_address', 'receiver_name', 'receiver_address',
        'direct_extraction_applied',  # New flag to show which records were corrected
        
        # Surcharges with proper triple format
        'fuel_surcharge', 'fuel_surcharge_published', 'fuel_surcharge_incentive', 'fuel_surcharge_billed',
        'residential_surcharge', 'residential_surcharge_published', 'residential_surcharge_incentive', 'residential_surcharge_billed',
        'delivery_area_surcharge', 'delivery_area_surcharge_published', 'delivery_area_surcharge_incentive', 'delivery_area_surcharge_billed',
        'large_package_surcharge', 'large_package_surcharge_published', 'large_package_surcharge_incentive', 'large_package_surcharge_billed',
        'additional_handling', 'additional_handling_published', 'additional_handling_incentive', 'additional_handling_billed',
        'saturday_delivery', 'saturday_delivery_published', 'saturday_delivery_incentive', 'saturday_delivery_billed',
        'signature_required', 'signature_required_published', 'signature_required_incentive', 'signature_required_billed',
        'adult_signature_required', 'adult_signature_required_published', 'adult_signature_required_incentive', 'adult_signature_required_billed',
        'address_correction', 'address_correction_published', 'address_correction_incentive', 'address_correction_billed',
        'over_maximum_limits', 'over_maximum_limits_published', 'over_maximum_limits_incentive', 'over_maximum_limits_billed',
        'peak_surcharge', 'peak_surcharge_published', 'peak_surcharge_incentive', 'peak_surcharge_billed',
        
        # Totals fields
        'line_total', 'line_total_published', 'line_total_incentive', 'line_total_billed',
        
        # Additional fields
        'dimensions', 'customer_weight', 'message_codes', 'number_of_packages',
        'control_id', 'shipped_from', 'bill_to', 'due_date', 'origin_zip',
        'billable_weight', 'dimensional_weight', 'package_type', 'net_charge',
        
        # Reference fields
        'first_reference', 'second_reference', 'third_reference', 'purchase_order',
        'invoice_reference', 'user_id',
        
        # Extended fields
        'cod_amount', 'declared_value', 'cod_surcharge', 'declared_value_charge',
        'delivery_date', 'commit_time', 'shipper_account', 'third_party_account',
        'hazmat_surcharge', 'dry_ice_surcharge', 'carbon_neutral', 'quantum_view',
        'ups_premium_care', 'missing_pld_fee'
    ]
    
    total_shipments = len(shipments)
    
    # Create rows for each invoice group with proper headings
    for invoice_num, invoice_shipments in invoice_groups.items():
        
        # Invoice group header with actual invoice number
        header_row = {
            'ROW_TYPE': 'INVOICE_HEADER',
            'INVOICE_GROUP_HEADER': f'Invoice: {invoice_num}',
            'SHIPMENT_INDEX': '',
            'SHIPMENT_COUNT': len(invoice_shipments),
            'TOTAL_SHIPMENTS': f'Total Shipments: {len(invoice_shipments)}',
            'invoice_number': invoice_num,
            'tracking_number': '',
            'account_number': invoice_shipments[0].get('account_number', ''),
            'invoice_date': invoice_shipments[0].get('invoice_date', ''),
            'processing_type': 'Matrix + Direct Field Extraction'
        }
        
        # Fill remaining columns with empty values for header
        for col in column_order:
            if col not in header_row:
                header_row[col] = ''
        
        excel_rows.append(header_row)
        
        # Shipment rows with corrected address data
        for shipment_idx, shipment in enumerate(invoice_shipments, 1):
            shipment_row = {
                'ROW_TYPE': f'Shipment {shipment_idx}',
                'INVOICE_GROUP_HEADER': '',
                'SHIPMENT_INDEX': shipment_idx,
                'SHIPMENT_COUNT': len(invoice_shipments),
                'TOTAL_SHIPMENTS': '',
                'invoice_number': invoice_num,
                'tracking_number': shipment.get('tracking_number', ''),
                'account_number': shipment.get('account_number', ''),
                'invoice_date': shipment.get('invoice_date', ''),
                'destination_zip': shipment.get('destination_zip', ''),
                'page_number': shipment.get('page_number', ''),
                'invoice_group': shipment.get('invoice_group', ''),
                'processing_type': 'Matrix + Direct Extraction',
                
                # Shipment details with proper formatting
                'weight': format_weight(shipment.get('weight')),
                'zone': shipment.get('zone', ''),
                'service_type': shipment.get('service_type', ''),
                'published_charge': format_currency(shipment.get('published_charge')),
                'incentive_credit': format_currency(shipment.get('incentive_credit')),
                'billed_charge': format_currency(shipment.get('billed_charge')),
                'shipment_date': shipment.get('shipment_date', ''),
                'pickup_date': shipment.get('pickup_date', ''),
                
                # CORRECTED - Address information from direct extraction
                'sender_name': shipment.get('sender_name', ''),  # Now corrected with direct extraction
                'sender_address': shipment.get('sender_address', ''),  # Now corrected with direct extraction
                'receiver_name': shipment.get('receiver_name', ''),  # Now corrected with direct extraction
                'receiver_address': shipment.get('receiver_address', ''),  # Now corrected with direct extraction
                'direct_extraction_applied': 'Yes' if shipment.get('direct_extraction_applied') else 'No',
                
                # Enhanced surcharge formatting
                'fuel_surcharge': format_surcharge_triple(shipment, 'fuel_surcharge'),
                'fuel_surcharge_published': format_currency(shipment.get('fuel_surcharge_published')),
                'fuel_surcharge_incentive': format_currency(shipment.get('fuel_surcharge_incentive')),
                'fuel_surcharge_billed': format_currency(shipment.get('fuel_surcharge_billed')),
                
                'residential_surcharge': format_surcharge_triple(shipment, 'residential_surcharge'),
                'residential_surcharge_published': format_currency(shipment.get('residential_surcharge_published')),
                'residential_surcharge_incentive': format_currency(shipment.get('residential_surcharge_incentive')),
                'residential_surcharge_billed': format_currency(shipment.get('residential_surcharge_billed')),
                
                'delivery_area_surcharge': format_surcharge_triple(shipment, 'delivery_area_surcharge'),
                'delivery_area_surcharge_published': format_currency(shipment.get('delivery_area_surcharge_published')),
                'delivery_area_surcharge_incentive': format_currency(shipment.get('delivery_area_surcharge_incentive')),
                'delivery_area_surcharge_billed': format_currency(shipment.get('delivery_area_surcharge_billed')),
                
                # Corrected totals calculation
                'line_total': format_surcharge_triple(shipment, 'line_total'),
                'line_total_published': format_currency(shipment.get('line_total_published')),
                'line_total_incentive': format_currency(shipment.get('line_total_incentive')),
                'line_total_billed': format_currency(shipment.get('line_total_billed')),
                
                # Other important fields
                'dimensions': shipment.get('dimensions', ''),
                'customer_weight': format_weight(shipment.get('customer_weight')),
                'message_codes': shipment.get('message_codes', ''),
                'control_id': shipment.get('control_id', ''),
                'shipped_from': shipment.get('shipped_from', ''),
                
                # Reference fields
                'first_reference': shipment.get('first_reference', ''),
                'second_reference': shipment.get('second_reference', ''),
                'third_reference': shipment.get('third_reference', ''),
                'user_id': shipment.get('user_id', ''),
                'purchase_order': shipment.get('purchase_order', '')
            }
            
            # Fill remaining columns with proper formatting
            for col in column_order:
                if col not in shipment_row:
                    value = shipment.get(col, '')
                    if isinstance(value, dict) and 'published' in value:
                        # Handle currency triple objects
                        shipment_row[col] = format_surcharge_triple(shipment, col)
                    elif col.endswith(('_charge', '_credit', '_billed', '_published', '_amount')):
                        shipment_row[col] = format_currency(value)
                    elif col.endswith('_weight'):
                        shipment_row[col] = format_weight(value)
                    else:
                        shipment_row[col] = value if value is not None else ''
            
            excel_rows.append(shipment_row)
    
    # Create DataFrame with proper column order
    df = pd.DataFrame(excel_rows)
    
    # Ensure all columns exist
    for col in column_order:
        if col not in df.columns:
            df[col] = ''
    
    df = df[column_order]
    
    # Save to Excel with enhanced formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='UPS_Invoice_Matrix_Corrected', index=False)
        
        # Apply enhanced formatting
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            workbook = writer.book
            worksheet = writer.sheets['UPS_Invoice_Matrix_Corrected']
            
            # Enhanced styles
            # Header styles
            header_font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
            header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Invoice header styles  
            invoice_header_font = Font(bold=True, size=12, color='1F4788', name='Calibri')
            invoice_header_fill = PatternFill(start_color='E8F1FF', end_color='E8F1FF', fill_type='solid')
            invoice_header_alignment = Alignment(horizontal='left', vertical='center')
            
            # Shipment styles
            shipment_font = Font(size=10, name='Calibri')
            shipment_fill_odd = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            shipment_fill_even = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            # Currency styles
            currency_font = Font(size=10, color='006600', name='Calibri')
            
            # Direct extraction applied styles (highlight corrected fields)
            corrected_font = Font(size=10, color='0066CC', name='Calibri', bold=True)
            corrected_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
            
            # Border styles
            thin_border = Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0')
            )
            
            # Apply column headers formatting
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Apply row-based formatting
            for row_idx in range(2, len(df) + 2):
                row_type = worksheet.cell(row=row_idx, column=1).value
                direct_extraction_applied = worksheet.cell(row=row_idx, column=df.columns.get_loc('direct_extraction_applied') + 1).value
                
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    column_name = df.columns[col_idx - 1]
                    
                    if isinstance(row_type, str):
                        if row_type == 'INVOICE_HEADER':
                            cell.font = invoice_header_font
                            cell.fill = invoice_header_fill
                            if col_idx <= 5:
                                cell.alignment = invoice_header_alignment
                        elif 'Shipment ' in row_type:
                            # Highlight corrected address fields
                            if (direct_extraction_applied == 'Yes' and 
                                column_name in ['sender_name', 'sender_address', 'receiver_name', 'receiver_address']):
                                cell.font = corrected_font
                                cell.fill = corrected_fill
                            else:
                                cell.font = shipment_font
                                if row_idx % 2 == 0:
                                    cell.fill = shipment_fill_even
                                else:
                                    cell.fill = shipment_fill_odd
                            
                            # Format currency columns
                            if cell.value and isinstance(cell.value, str) and cell.value.startswith(''):
                                if not (direct_extraction_applied == 'Yes' and 
                                       column_name in ['sender_name', 'sender_address', 'receiver_name', 'receiver_address']):
                                    cell.font = currency_font
            
            # Enhanced column width adjustment
            column_widths = {
                'ROW_TYPE': 15,
                'INVOICE_GROUP_HEADER': 25,
                'SHIPMENT_INDEX': 8,
                'SHIPMENT_COUNT': 8,
                'TOTAL_SHIPMENTS': 18,
                'invoice_number': 25,
                'tracking_number': 22,
                'account_number': 15,
                'invoice_date': 12,
                'service_type': 20,
                'destination_zip': 10,
                'published_charge': 12,
                'incentive_credit': 12,
                'billed_charge': 12,
                'sender_name': 30,  # Increased for corrected names
                'sender_address': 45,  # Increased for corrected addresses  
                'receiver_name': 30,  # Increased for corrected names
                'receiver_address': 45,  # Increased for corrected addresses
                'direct_extraction_applied': 15,  # New column
                'fuel_surcharge': 40,
                'residential_surcharge': 40,
                'delivery_area_surcharge': 40,
                'line_total': 40,
                'dimensions': 18,
                'message_codes': 12,
                'first_reference': 15,
                'second_reference': 15,
                'user_id': 15
            }
            
            for col_idx, column in enumerate(worksheet.columns, 1):
                column_letter = get_column_letter(col_idx)
                column_name = df.columns[col_idx - 1] if col_idx <= len(df.columns) else ''
                
                if column_name in column_widths:
                    width = column_widths[column_name]
                elif '_published' in column_name or '_incentive' in column_name or '_billed' in column_name:
                    width = 12
                elif '_address' in column_name:
                    width = 45  # Increased for better visibility
                elif '_reference' in column_name or 'user_id' in column_name:
                    width = 15
                elif '_surcharge' in column_name and not column_name.endswith(('_published', '_incentive', '_billed')):
                    width = 40
                elif column_name.startswith('line_total'):
                    width = 15
                else:
                    width = 12
                
                worksheet.column_dimensions[column_letter].width = width
            
            # Freeze panes for better navigation
            worksheet.freeze_panes = 'F2'
            
        except ImportError:
            logger.warning("Advanced Excel formatting not available - basic Excel created")
    
    logger.info(f"Enhanced Matrix Excel file with corrected address fields created: {len(invoice_groups)} invoices, {len(shipments)} shipments")

def format_currency(value):
    """Format currency value"""
    if value is None or value == '':
        return ''
    try:
        if isinstance(value, str) and value.startswith(''):
            return value
        return f"${float(value):.2f}" if value else ''
    except (ValueError, TypeError):
        return str(value) if value else ''

def format_weight(value):
    """Format weight value"""
    if value is None or value == '':
        return ''
    try:
        return f"{float(value)} lbs" if value else ''
    except (ValueError, TypeError):
        return str(value) if value else ''

def format_surcharge_triple(shipment, surcharge_name):
    """Format surcharge in the exact format: Published: $X.XX, Incentive: $Y.YY, Billed: $Z.ZZ"""
    published = shipment.get(f"{surcharge_name}_published")
    incentive = shipment.get(f"{surcharge_name}_incentive")
    billed = shipment.get(f"{surcharge_name}_billed")
    
    if not any([published, incentive, billed]):
        return ''
    
    parts = []
    if published is not None:
        try:
            parts.append(f"Published: ${float(published):.2f}")
        except (ValueError, TypeError):
            pass
    if incentive is not None:
        try:
            parts.append(f"Incentive: ${float(incentive):.2f}")
        except (ValueError, TypeError):
            pass
    if billed is not None:
        try:
            parts.append(f"Billed: ${float(billed):.2f}")
        except (ValueError, TypeError):
            pass
    
    return ', '.join(parts) if parts else ''

def generate_processing_statistics(shipments):
    """Generate detailed processing statistics with direct extraction metrics"""
    stats = {
        'total_shipments': len(shipments),
        'service_types': {},
        'zones': {},
        'total_charges': {'published': 0.0, 'incentive': 0.0, 'billed': 0.0},
        'field_coverage': {},
        'invoice_count': len(set(s.get('invoice_number') for s in shipments if s.get('invoice_number'))),
        'receiver_data_coverage': 0,
        'sender_data_coverage': 0,
        'direct_extraction_stats': {
            'total_corrected': 0,
            'sender_name_corrected': 0,
            'sender_address_corrected': 0,
            'receiver_name_corrected': 0,
            'receiver_address_corrected': 0
        }
    }
    
    # Analyze shipments
    receiver_count = 0
    sender_count = 0
    direct_extraction_count = 0
    
    for shipment in shipments:
        # Service types
        service = shipment.get('service_type', 'Unknown')
        stats['service_types'][service] = stats['service_types'].get(service, 0) + 1
        
        # Zones
        zone = shipment.get('zone', 'Unknown')
        stats['zones'][zone] = stats['zones'].get(zone, 0) + 1
        
        # Total charges using line totals
        line_total_pub = shipment.get('line_total_published')
        line_total_inc = shipment.get('line_total_incentive') 
        line_total_bill = shipment.get('line_total_billed')
        
        if isinstance(line_total_pub, (int, float)):
            stats['total_charges']['published'] += line_total_pub
        if isinstance(line_total_inc, (int, float)):
            stats['total_charges']['incentive'] += line_total_inc
        if isinstance(line_total_bill, (int, float)):
            stats['total_charges']['billed'] += line_total_bill
        
        # Address coverage
        if shipment.get('receiver_name') or shipment.get('receiver_address'):
            receiver_count += 1
        if shipment.get('sender_name') or shipment.get('sender_address'):
            sender_count += 1
        
        # Direct extraction statistics
        if shipment.get('direct_extraction_applied'):
            direct_extraction_count += 1
            stats['direct_extraction_stats']['total_corrected'] += 1
            
            if shipment.get('sender_name'):
                stats['direct_extraction_stats']['sender_name_corrected'] += 1
            if shipment.get('sender_address'):
                stats['direct_extraction_stats']['sender_address_corrected'] += 1
            if shipment.get('receiver_name'):
                stats['direct_extraction_stats']['receiver_name_corrected'] += 1
            if shipment.get('receiver_address'):
                stats['direct_extraction_stats']['receiver_address_corrected'] += 1
    
    stats['receiver_data_coverage'] = (receiver_count / len(shipments) * 100) if shipments else 0
    stats['sender_data_coverage'] = (sender_count / len(shipments) * 100) if shipments else 0
    stats['direct_extraction_coverage'] = (direct_extraction_count / len(shipments) * 100) if shipments else 0
    
    # Calculate field coverage
    all_possible_fields = [
        'tracking_number', 'service_type', 'destination_zip', 'weight', 'zone',
        'published_charge', 'incentive_credit', 'billed_charge',
        'line_total_published', 'line_total_incentive', 'line_total_billed',
        'residential_surcharge_published', 'fuel_surcharge_published',
        'receiver_name', 'receiver_address', 'sender_name', 'sender_address',
        'user_id', 'first_reference', 'second_reference'
    ]
    
    for field in all_possible_fields:
        populated = sum(1 for s in shipments if s.get(field) not in [None, '', 'Unknown'])
        stats['field_coverage'][field] = {
            'populated': populated,
            'percentage': (populated / len(shipments) * 100) if shipments else 0
        }
    
    return stats

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        logger.info("=== ENHANCED UPLOAD WITH DIRECT FIELD EXTRACTION ===")
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # if not allowed_file(file.filename):
        #     return jsonify({'error': 'Only PDF files are allowed'}), 400
        
        # Get session ID
        session_id = request.form.get('session_id')
        if not session_id:
            return jsonify({'error': 'Session ID required'}), 400
        
        logger.info(f"Processing file: {file.filename} for session: {session_id}")
        
        # Generate unique filenames
        file_id = str(uuid.uuid4())
        filename = secure_filename(file.filename)
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file_id}_{filename}")
        
        # Use .xlsx extension for Excel output
        excel_filename = f"{file_id}_ups_corrected_matrix.xlsx"
        excel_path = os.path.join(app.config['OUTPUT_FOLDER'], excel_filename)
        
        # Save uploaded file
        file.save(pdf_path)
        
        # Start processing in a separate thread
        def process_file():
            try:
                logger.info(f"Starting enhanced processing with direct extraction for session: {session_id}")
                process_invoice_with_progress(pdf_path, excel_path, session_id)
                
                # Clean up uploaded PDF
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
                
                # Remove session from active sessions
                if session_id in active_sessions:
                    del active_sessions[session_id]
                    
            except Exception as e:
                logger.error(f"Error in background processing: {e}", exc_info=True)
                # Clean up files on error
                try:
                    if os.path.exists(pdf_path):
                        os.remove(pdf_path)
                    if os.path.exists(excel_path):
                        os.remove(excel_path)
                except:
                    pass
                
                if session_id in active_sessions:
                    del active_sessions[session_id]
        
        thread = Thread(target=process_file)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'message': 'Enhanced processing started with direct field extraction for accurate address data',
            'download_filename': excel_filename,
            'processing_type': 'Matrix + Direct Field Extraction (Corrected Address Fields)'
        })
            
    except Exception as e:
        logger.error(f"Upload error: {e}", exc_info=True)
        return jsonify({'error': f'Upload failed: {str(e)}'}), 500

@app.route('/download/<filename>')
def download_file(filename):
    try:
        file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404
        
        return send_file(file_path, as_attachment=True, download_name=filename)
    
    except Exception as e:
        logger.error(f"Download error: {e}")
        return jsonify({'error': 'Download failed'}), 500

@socketio.on('connect')
def handle_connect():
    logger.info('=== CLIENT CONNECTED ===')
    emit('connected', {'data': 'Connected to Enhanced UPS Processing Server with Direct Field Extraction'})

@socketio.on('disconnect')
def handle_disconnect():
    logger.info('=== CLIENT DISCONNECTED ===')

@socketio.on('join_session')
def handle_join_session(data):
    session_id = data.get('session_id')
    if session_id:
        join_room(session_id)
        active_sessions[session_id] = {'connected_at': time.time()}
        logger.info(f'=== CLIENT JOINED SESSION: {session_id} ===')
        emit('session_joined', {'session_id': session_id, 'processing_type': 'Matrix + Direct Field Extraction'})

@app.errorhandler(413)
def too_large(e):
    return jsonify({'error': 'File too large. Maximum size is 50MB'}), 413

@app.errorhandler(500)
def internal_error(e):
    logger.error(f"Internal server error: {e}")
    return jsonify({'error': 'Internal server error'}), 500

if __name__ == '__main__':
    logger.info("=== STARTING ENHANCED UPS PROCESSING SERVER WITH DIRECT FIELD EXTRACTION ===")
    logger.info("New Features:")
    logger.info("1. ✓ Direct PDF text extraction for 5 key fields")
    logger.info("2. ✓ Accurate tracking number, sender/receiver name & address extraction")
    logger.info("3. ✓ Field replacement based on tracking number matching")
    logger.info("4. ✓ Enhanced Excel with corrected address fields highlighted")
    logger.info("5. ✓ Dual extraction approach: Matrix + Direct field correction")
    socketio.run(app, debug=True, host='0.0.0.0', port=5000)
 